from django.shortcuts import render
from sap_conn import sap_conn
import pandas as pd  # 做資料分析的套件
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, Alignment
from openpyxl.utils import get_column_letter
# Create your views here.
from django.template import context
from django.shortcuts import HttpResponse

def index(request):
    # context['object_list'] = article.objects.filter( title__icontains=request.GET.get('search'))
    # return render( request, "encyclopedia/article_detail.html", context )
    if request.method == 'POST':
        from_date = request.POST.get('from_date')  # date variable for selection-screen 2
        to_date = request.POST.get('to_date')  # date variable for selection-screen 3.
        item_name = request.POST.get('item_name')  # 物料號 for selection-screen 4
        print(from_date)  # string 2021-08-31
        print(to_date)
        print(item_name)  # string

        from_date_str = from_date  # date variable for selection-screen 2.
        to_date_str = to_date  # date variable for selection-screen 3.
        item_name = item_name  # 物料號 for selection-screen 4
        # create session
        request.session['from_date_str'] = from_date_str
        request.session['to_date_str'] = to_date_str
        request.session['item_name'] = item_name

        from_date = pd.to_datetime( from_date_str, format='%Y-%m-%d' )
        to_date = pd.to_datetime( to_date_str, format='%Y-%m-%d' )
        item_name = item_name.strip()
        df = sap_conn.to_dataframe()
        df['date'] = pd.to_datetime(df.date)
        df['year'] = pd.DatetimeIndex(df.date).year  # int64
        df['month'] = pd.DatetimeIndex(df.date).month  # int64
        df['day'] = pd.DatetimeIndex(df.date).day  # int64
        print(df.info())
        prix_tax_incl = pd.to_numeric( df['prix_sans_tax']) * 100 * 1.05
        df['prix_tax_incl'] = np.floor(prix_tax_incl + 0.5).astype(
            np.int64)  # astype(np.int64)把原先float的trailing 0除掉
        df['date'] = pd.to_datetime(df['date'])
        if item_name == '':
            for_count = df.query('date >= @from_date & date <= @to_date')  # _for_count用來做累計訂單數統計
        else:
            for_count = df.query('date >= @from_date & date <= @to_date & item == @item_name')
        print('\n---同一月份的同一個物料的累積quantity---(統計概念)')
        item_groups = df.groupby(by=['item', 'date', 'description'], as_index=False)[
            'quantity', 'prix_tax_incl'].sum()
        if item_name == '':
            april_all_items = item_groups.query( 'date >= @from_date & date <= @to_date' )
        else:
            april_all_items = item_groups.query(
                'date >= @from_date & date <= @to_date & item == @item_name' )  # 撈出所有有此物料的row沒有做加總
        if april_all_items.empty:
            return render(request, "selection_screen_yeezen.html", {
                'no_data':'查無資料',
                'item_name': request.session['item_name'],
                'from_date': request.session['from_date_str'],
                'to_date': request.session['to_date_str']
            })
            del request.session['no_data']
            del request.session['item_name']
            del request.session['from_date']
            del request.session['to_date']
        else:
            print( 'april_all_items QUERY date, item_name\n', april_all_items )  # 所有在日期區間的資料
            value_counts = for_count['item'].value_counts()  # return list
            df_value_counts = pd.DataFrame( value_counts )  # put list to dataframe
            df_value_counts = df_value_counts.reset_index()
            df_value_counts.columns = ['物料號碼', '累計訂單數']  # change column names
            print( df_value_counts )  # ['物料號碼', '累計訂單數']
            print( from_date_str + '_' + to_date_str + '的所有物料累績: \n', april_all_items )
            print( april_all_items.shape )  # 308 rows
            april_all_items = april_all_items.groupby( by=['item', 'description'], as_index=False )[
                'quantity', 'prix_tax_incl'].sum()
            print( '用item, description 做SUM: \n', april_all_items )  # 264    OZCOU 最後一筆資料沒有加總正確
            presented_to_excel_df = april_all_items[['item', 'description', 'quantity', 'prix_tax_incl']]  # year, day不需要相加
            print( 'presented_to_excel_df: \n', presented_to_excel_df )
            # sort values by cols
            print( '---sort values by cols----(數據呈現)' )
            sorted_by_item_df = df.sort_values( ['date', 'item'] )  # 日期由小到大排序、
            item = sorted_by_item_df['item']
            sorted_by_item_df = sorted_by_item_df[
                ['date', 'item', 'description', 'quantity', 'prix_sans_tax', 'shipment_no', 'prix_tax_incl']]
            print( sorted_by_item_df.head(20))
            # 寫進excel
            if item_name != '':
                writer = pd.ExcelWriter(item_name + '_' + from_date_str + '_' + to_date_str + '_' + '怡仁訂購統計.xlsx')
            else:
                writer = pd.ExcelWriter(from_date_str + '_' + to_date_str + '_' + '怡仁訂購統計.xlsx')
            presented_to_excel_df.to_excel(writer, sheet_name='怡仁訂購統計', header=None, index=False,
                                            startrow=1)  # names=['物料號','中英文名', '累計訂購數量', '累計含稅金額', 'how_many(當月買幾次)']
            sorted_by_item_df.to_excel(writer, sheet_name='總表', header=True, index=False)  # , startcol=7, startrow=1
            df_value_counts.to_excel(writer, sheet_name='怡仁訂購統計', header=None, index=False, startcol=6, startrow=1)
            writer.save()

            print( 'DataFrame is written successfully to Excel File.' )
            if item_name != '':
                wb = load_workbook(item_name + '_' + from_date_str + '_' + to_date_str + '_' + '怡仁訂購統計.xlsx')
            else:
                wb = load_workbook(from_date_str + '_' + to_date_str + '_' + '怡仁訂購統計.xlsx')

            sheet = wb['怡仁訂購統計']

            sheet['A1'] = '物料號'
            sheet['B1'] = '中英文名'
            sheet['C1'] = '累計訂購數量'
            sheet['D1'] = '累計訂購金額(含稅)'
            sheet['E1'] = '累計訂單數'  # (當月買幾次)
            sheet.freeze_panes = sheet['A2']
            # ['date', 'item', 'quantity', 'prix_sans_tax', 'shipment_no', 'prix_tax_incl']
            sheet2 = wb['總表']
            sheet2['A1'] = '日期'
            sheet2['B1'] = '物料號'
            sheet2['C1'] = '中英文名'
            sheet2['D1'] = '數量'
            sheet2['E1'] = '未稅金額'
            sheet2['F1'] = '出貨單號'
            sheet2['G1'] = '含稅金額'
            sheet2.freeze_panes = sheet2['A2']

            right_align = Alignment(horizontal='right', vertical='top')  # , wrap_text=False
            font = Font( name='Arial', size=11 )
            date_style = NamedStyle(name='datetime', number_format='YYYY/MM/DD')
            for row in sheet2['A2:A' + str( sheet2.max_row )]:
                for cell in row:
                    cell.style = date_style
            for row in sheet2['A1:' + get_column_letter(sheet2.max_column) + str(sheet2.max_row)]:
                for cell in row:
                    cell.alignment = right_align
                    cell.font = font
            for row in sheet['A1:' + get_column_letter(sheet.max_column) + str(sheet.max_row)]:
                for cell in row:
                    cell.alignment = right_align
                    cell.font = font

            print( '**********************' )
            for i in range( 1, sheet.max_row + 1 ):
                for j in range( 1, sheet.max_row + 1 ):
                    # print('G' + str(j), sheet['G' + str(j)].value)
                    # print('Ax' + str(i), sheet['A' + str(i)].value)
                    if sheet['G' + str( j )].value == sheet['A' + str( i )].value:
                        sheet['E' + str( i )].value = sheet['H' + str( j )].value
                        sheet['G' + str( j )].value = ''
                        sheet['H' + str( j )].value = ''
                        # print('Ai', i)
                        # print('Gj', j)
            # 訂購統計
            sheet.column_dimensions['B'].width = 55
            sheet.column_dimensions['A'].width = 15
            sheet.column_dimensions['C'].width = 18
            sheet.column_dimensions['D'].width = 23
            sheet.column_dimensions['E'].width = 18
            # 總表
            sheet2.column_dimensions['C'].width = 55
            sheet2.column_dimensions['A'].width = 15
            sheet2.column_dimensions['B'].width = 15
            sheet2.column_dimensions['D'].width = 18
            sheet2.column_dimensions['E'].width = 18
            sheet2.column_dimensions['F'].width = 18
            sheet2.column_dimensions['G'].width = 18
            if item_name != '':
                wb.save(item_name + '_' + from_date_str + '_' + to_date_str + '_' + '怡仁訂購統計.xlsx' )
            else:
                wb.save(from_date_str + '_' + to_date_str + '_' + '怡仁訂購統計.xlsx')
            return render(request, "selection_screen_yeezen.html", {
                               'item_name': request.session['item_name'],
                               'from_date': request.session['from_date_str'],
                                'to_date': request.session['to_date_str']
                               })

    else:
        return render(request, "selection_screen_yeezen.html")

def download(request):
    from_date_str = request.session['from_date_str']
    to_date_str = request.session['to_date_str']
    item_name = request.session['item_name']
    if item_name != '':
        file = open(item_name + '_' + from_date_str + '_' + to_date_str + '_' + '怡仁訂購統計.xlsx', 'rb')
        response = HttpResponse(file)
        filename = item_name + '_' + from_date_str + '_' + to_date_str + '.xlsx'
        response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'  # 設定頭資訊，告訴瀏覽器這是個檔案
        # response['Content-Disposition'] = 'attachment; filename= "%s"' % filename
        # response['Content-Disposition'] = 'attachment; filename= Consistency Report.xlsx'
        response['Content-Disposition'] = 'attachment; filename=  "%s"' % filename
    else:
        file = open(from_date_str + '_' + to_date_str + '_' + '怡仁訂購統計.xlsx', 'rb')
        response = HttpResponse(file)
        response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet; charset=utf-8'
        filename = from_date_str + '_' + to_date_str + '.xlsx'
        response['Content-Disposition'] = 'attachment; filename= "%s"' % filename
    return response